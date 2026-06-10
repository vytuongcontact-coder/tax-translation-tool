"""
build_vi_en_glossary.py
=======================
Script tạo Glossary VI→EN (formatted) từ Glossary_EN_VI_2_200526.xlsx.

Logic:
  1. Đọc file EN→VI gốc (558 rows)
  2. Drop trùng (EN, VI) pair → normalize VI về Title Case → pick 1 EN duy nhất / VI
  3. Swap: VI (source) → EN (target)
  4. Sinh 8 biến thể: 4 styles × 2 (italic / non-italic)
  5. Xuất Excel đẹp (3 sheets)
"""

import pandas as pd
import openpyxl
import unicodedata
import os
import sys

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ─────────────────────────────────────────────────────────────────────────────
# Palette
# ─────────────────────────────────────────────────────────────────────────────

COL_HEADER_BG  = "1F4E79"
COL_HEADER_FG  = "FFFFFF"
# 8 colors: 4 styles × italic / non-italic
COL_LOWER_BG      = "E8F5E9"   # soft green
COL_LOWER_ITAL_BG = "D5F5E3"   # deeper green (italic)
COL_ORIG_BG       = "F1F8E9"   # original VI
COL_ORIG_ITAL_BG  = "E8F8F5"   # slightly deeper (italic)
COL_NODIA_BG      = "C8E6C9"   # mid green
COL_NODIA_ITAL_BG = "A9D6BE"   # deeper green (italic)
COL_UPPER_BG      = "AED581"   # stronger green
COL_UPPER_ITAL_BG = "8BC34A"   # deeper green (italic)
COL_TARGET_BG     = "FFF9C4"   # soft yellow – EN translation

THIN = Side(style="thin",   color="A5D6A7")
MED  = Side(style="medium", color="1B5E20")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_MED  = Border(left=MED,  right=MED,  top=MED,  bottom=MED)


# ─────────────────────────────────────────────────────────────────────────────
# Vietnamese helpers
# ─────────────────────────────────────────────────────────────────────────────

def strip_diacritics(text: str) -> str:
    nfd = unicodedata.normalize("NFD", text)
    return "".join(c for c in nfd if unicodedata.category(c) != "Mn").lower()


def _vi_case_variations(vi_term: str):
    """
    Sinh đúng 8 biến thể: 4 styles × 2 (plain / italic).
    Mỗi entry đều có đủ 8 dòng.
    """
    lower_text = vi_term.lower()
    no_dia_text = strip_diacritics(vi_term)
    upper_text = vi_term.upper()

    yield ("lower",        lower_text,   False)   # plain
    yield ("lower",        lower_text,   True)    # italic
    yield ("original",     vi_term,      False)   # plain
    yield ("original",     vi_term,      True)    # italic
    yield ("no_diacritic", no_dia_text, False)   # plain
    yield ("no_diacritic", no_dia_text, True)     # italic
    yield ("upper",        upper_text,   False)   # plain
    yield ("upper",        upper_text,   True)    # italic


# ─────────────────────────────────────────────────────────────────────────────
# Style factories
# ─────────────────────────────────────────────────────────────────────────────

def header_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Calibri", bold=True, color=COL_HEADER_FG, size=11)
    c.fill      = PatternFill("solid", fgColor=COL_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = BORDER_MED
    return c


def data_cell(ws, row, col, value, bg, bold=False, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Calibri", bold=bold, italic=italic, color="1A1A2E", size=10)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    c.border    = BORDER_THIN
    return c


def translation_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Calibri", color="1A1A2E", size=10)
    c.fill      = PatternFill("solid", fgColor=COL_TARGET_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    c.border    = BORDER_THIN
    return c


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

BG_MAP = {
    ("lower",        False): COL_LOWER_BG,
    ("lower",        True):  COL_LOWER_ITAL_BG,
    ("original",     False): COL_ORIG_BG,
    ("original",     True):  COL_ORIG_ITAL_BG,
    ("no_diacritic", False): COL_NODIA_BG,
    ("no_diacritic", True):  COL_NODIA_ITAL_BG,
    ("upper",        False): COL_UPPER_BG,
    ("upper",        True):  COL_UPPER_ITAL_BG,
}

LEGEND_LABELS = {
    ("lower",        False): ("lower",              COL_LOWER_BG,     "chữ thường – nguyên tắc giao dịch độc lập"),
    ("lower",        True):  ("lower (italic)",     COL_LOWER_ITAL_BG,"chữ thường – in nghiêng"),
    ("original",     False): ("original",            COL_ORIG_BG,       "giữ nguyên – Nguyên Tắc Giao Dịch Độc Lập"),
    ("original",     True):  ("original (italic)",   COL_ORIG_ITAL_BG, "giữ nguyên – in nghiêng"),
    ("no_diacritic", False): ("no_diacritic",        COL_NODIA_BG,     "bỏ dấu – nguyen tac giao dich doc lap"),
    ("no_diacritic", True):  ("no_diacritic (italic)", COL_NODIA_ITAL_BG, "bỏ dấu – in nghiêng"),
    ("upper",        False): ("upper",              COL_UPPER_BG,     "IN HOA – NGUYÊN TẮC GIAO DỊCH ĐỘC LẬP"),
    ("upper",        True):  ("upper (italic)",     COL_UPPER_ITAL_BG,"IN HOA – in nghiêng"),
}


def build(source_path: str, output_path: str):
    # ── Step 1: Read & deduplicate ─────────────────────────────────────────
    df = pd.read_excel(source_path)
    df.columns = [c.strip() for c in df.columns]
    df["Source"] = df["Source"].str.strip()
    df["Target"] = df["Target"].str.strip()

    # Drop exact (EN, VI) duplicates
    pairs = df[["Source", "Target"]].drop_duplicates().reset_index(drop=True)

    # ── Step 2: Canonical VI + pick 1 EN per VI ────────────────────────────
    def pick_canonical_en(grp):
        rows = grp.sort_values("Source").to_dict("records")
        for r in rows:
            en = str(r["Source"]).strip()
            if en.title() == en and not en.isupper():
                return en
        return rows[0]["Source"].strip()

    pairs["vi_norm"] = pairs["Target"].apply(
        lambda v: " ".join(w.capitalize() for w in str(v).split())
    )

    canonical = (
        pairs.groupby("vi_norm", sort=False, group_keys=False)
        .apply(lambda g: pick_canonical_en(g), include_groups=False)
        .reset_index()
    )
    canonical.columns = ["vi", "en"]

    print(f"Source: {len(df)} rows → {len(pairs)} unique pairs → {len(canonical)} VI→EN entries")

    # ── Step 3: Build 8-style rows ──────────────────────────────────────────
    rows = []   # (vi_canonical, style, variant, italic, en)
    for _, row in canonical.iterrows():
        vi_canon = row["vi"]
        en_trans = row["en"]
        if not vi_canon or not en_trans:
            continue
        for style_name, variant, italic in _vi_case_variations(vi_canon):
            rows.append((vi_canon, style_name, variant, italic, en_trans))

    print(f"Formatted rows: {len(rows)} (8 per term)")

    # Validate
    entry_count = {}
    for (vi, style, _, italic, _) in rows:
        key = vi
        entry_count[key] = entry_count.get(key, 0) + 1
    bad = {v: c for v, c in entry_count.items() if c != 8}
    if bad:
        print(f"WARNING: {len(bad)} entries have != 8 rows")
    else:
        print(f"All {len(canonical)} entries have exactly 8 rows.")

    def style_label(style: str, italic: bool) -> str:
        if italic:
            return f"{style} (italic)"
        return style

    # ── Step 4: Write workbook ────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Glossary (VI→EN Formatted)"
    ws.freeze_panes = "A2"

    header_cell(ws, 1, 1, "Case Style")
    header_cell(ws, 1, 2, "Keyword (VI)")
    header_cell(ws, 1, 3, "Translation (EN)")

    for i, (vi_canon, style, variant, italic, en_text) in enumerate(rows, start=2):
        bg = BG_MAP.get((style, italic), "F1F8E9")
        label = style_label(style, italic)

        c1 = ws.cell(row=i, column=1, value=label)
        c1.font      = Font(name="Calibri", italic=True, color="5A6A7A", size=9)
        c1.fill      = PatternFill("solid", fgColor=bg)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.border    = BORDER_THIN

        data_cell(ws, i, 2, variant, bg,
                  bold=(style == "upper"), italic=italic)
        translation_cell(ws, i, 3, en_text)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 55
    ws.row_dimensions[1].height = 30
    for r in range(2, len(rows) + 2):
        ws.row_dimensions[r].height = 18
    ws.auto_filter.ref = f"A1:C{len(rows) + 1}"

    # ── Sheet 2: Legend ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Legend")
    header_cell(ws2, 1, 1, "Style Name")
    header_cell(ws2, 1, 2, "Example")
    header_cell(ws2, 1, 3, "Description")
    for i, ((style, italic), (_, bg, example)) in enumerate(LEGEND_LABELS.items(), start=2):
        for col, val in enumerate([style, example, "Biến thể cho cột tiếng Việt"], start=1):
            c = ws2.cell(row=i, column=col, value=val)
            c.font      = Font(name="Calibri", size=10, color="1A1A2E")
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            c.border    = BORDER_THIN
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 38

    # ── Sheet 3: Raw pairs ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Raw Pairs (VI→EN)")
    header_cell(ws3, 1, 1, "Keyword (VI)")
    header_cell(ws3, 1, 2, "Translation (EN)")
    for i, (_, row) in enumerate(canonical.iterrows(), start=2):
        ws3.cell(row=i, column=1, value=row["vi"])
        ws3.cell(row=i, column=2, value=row["en"])
    ws3.column_dimensions["A"].width = 60
    ws3.column_dimensions["B"].width = 60
    ws3.auto_filter.ref = f"A1:B{len(canonical) + 1}"

    wb.save(output_path)
    print(f"Saved: {output_path}")
    return len(rows), len(canonical)


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    base = "/Users/tuongvy/Documents/Tax Translation/web_tool/uploads"
    src  = "/Users/tuongvy/Documents/Tax Translation/Glossary_EN_VI_2_200526.xlsx"
    out  = os.path.join(base, "Glossary_VI_EN_formatted.xlsx")

    if len(sys.argv) >= 3:
        src = sys.argv[1]
        out = sys.argv[2]

    rows, pairs = build(src, out)
    print(f"Done — {rows} formatted rows from {pairs} unique glossary entries")
